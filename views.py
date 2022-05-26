#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2022-03-12 21:37
# @Author  : pan
# @Site    :
# @File    : work.py
# @Software: PyCharm
# @Description: 丰润财务日常工作脚本，包含文件分类、客户文件夹操作、电子发票操作、邮件操作等
import datetime
import os
import re
import shutil
from collections import namedtuple
from pathlib import Path

import openpyxl
import xlwings as xw
from openpyxl.styles import Alignment, Font, colors
from peewee import fn
from setuptools.msvc import winreg

from config import CUS_FILE_RANGE
from fastapi_models import Customer
from models import SocialHouse, User, Customer


def get_differ_month(start_date, end_date):
    """
    [计算连个日期之间相差的月份数]

    Args:
        start_date ([datetime]): [起始时间]
        end_date ([datetime]): [截止时间]

    Returns:
        [int]: [返回月份差的整数]
    """
    year_end = int(end_date.year)
    month_end = int(end_date.month)
    year_start = int(start_date.year)
    month_start = int(start_date.month)
    return (year_start - year_end) * 12 + (month_start - month_end) + 1


def read_from_excel_by_xw(file, table=None):
    """
    :param file: excel文件的路径全称，例如：r'I:\2、work\1、FullRain_System\客户查询系统.xlsx'
    :param table: excel表格名称
    :return: 返回一个列表
    """
    app = xw.App(visible=False, add_book=False)
    wb = app.books.open(file)
    try:
        sht = wb.sheets[table] if table else wb.sheets.active
        shape = sht.used_range.shape
        return sht[:shape[0], :shape[1]].value
    except Exception as e:
        print(e)
        return []
    finally:
        wb.close()
        app.quit()


class FileHundle:

    def file_range_by_suffix(self, path, file_path):

        """
        func: 将文件按照后缀名分类存入对应文件夹下
        :param path: 待处理文件所在的根目录
        :param file_path:
        :return:
        """
        file = os.listdir(path)  # 列出当前文件夹的所有文件
        for f in file:  # 循环遍历每个文件
            folder_name = file_path + f.split(".")[-1]  # 以扩展名为名称的子文件夹
            if not os.path.exists(folder_name):  # 如果不存在该目录
                os.makedirs(folder_name)  # 先创建，再移动文件
            shutil.move(f, folder_name)

    def file_range_by_type(self, scr_path):
        """
        func:将文件按照文件类型存入对应文件夹下
        :param scr_path: 待分类文件所在的目录
        :return: 无
        """
        FILE_FORMATS = {
            "图片资料": [".jpeg", ".jpg", ".tiff", ".gif", ".bmp", ".png", ".bpg", "svg", ".heif", ".psd"],
            "文档资料": [".oxps", ".epub", ".pages", ".docx", ".doc", ".fdf", ".ods", ".odt", ".pwi", ".xsn",
                     ".xps", ".dotx", ".docm", ".dox", ".rvg", ".rtf", ".rtfd", ".wpd", ".xls", ".xlsx",
                     ".xlsm", ".ppt", ".pptx", ".csv", ".pdf", ".md", ".xmind"],
            "视频文件": [".avi", ".flv", ".wmv", ".mov", ".mp4", ".webm", ".vob", ".mng", ".qt", ".mpg", ".mpeg", ".3gp",
                     ".mkv"],
            "压缩文件": [".a", ".ar", ".cpio", ".iso", ".tar", ".gz", ".rz", ".7z", ".dmg", ".rar", ".xar", ".zip"],
            "程序文件": [".exe", ".bat", ".lnk", ".js", ".dll", ".db", ".py", ".html5", ".html", ".htm", ".xhtml",
                     ".cpp", ".java", ".css", ".sql", ".msi"],
            "网页文件": ['.html', '.xml', '.mhtml', '.html'],
            "音频文件": [".aac", ".aa", ".aac", ".dvf", ".m4a", ".m4b", ".m4p", ".mp3", ".msv", ".ogg", ".oga",
                     ".raw", ".vox", ".wav", ".wma"],
        }
        for my_file in Path(scr_path).glob("**/*"):
            if my_file.is_dir():
                # 用continue就跳过了文件夹
                continue
            file_path = Path(scr_path + '\\' + my_file.name)  # 拼接形成文件
            lower_file_path = file_path.suffix.lower()  # 后缀转化成小写
            for my_key, value in FILE_FORMATS.items():
                if lower_file_path in value:  # 如果后缀名在上面定义的
                    directory_path = Path(scr_path + '\\' + my_key)

                    # 如果文件夹不存在，则根据定义建立文件夹
                    directory_path.mkdir(exist_ok=True)
                    new_name = directory_path.joinpath(my_file.name)
                    print(new_name)
                    shutil.move(my_file, new_name)
                    # file_path.replace(directory_path.joinpath(my_file.name))
        print('文件分类已结束！')

    def move_dzfp_pdf(self, all_files_list, dst_path):
        """
        func: 筛选出电子发票移动到目标目录
        :param all_files_list: 所有文件的列表
        :param dst_path: 电子发票移动目录
        :return:
        """
        for file_name in all_files_list:
            try:
                if file_name.endswith(".pdf"):  # 判断是否是PDF文件
                    height, width = self.get_pdf_size(file_name)
                    list1 = file_name.split('\\')
                    new_file_name = os.path.join(dst_path, list1[-1])
                    # 通过识别PDF第一页尺寸判断文件是否是电子发票
                    if 390 < height < 400 and 590 < width < 650:
                        print(f"这是一个电子发票：{file_name}，页面宽高比：{width / height}，移动到：{dst_path}")

                        shutil.move(file_name, new_file_name)  # 移动文件
                    else:
                        print(f"提示：非电子发票：{file_name}，页面宽高比：{width / height}")
                else:
                    continue
            except Exception as e:
                print(e)

    def get_pdf_size(self, filename):
        """
        func:获取PDF页面尺寸
        :param filename: pdf文件的名称
        """
        try:
            with open(filename, 'rb') as f:
                pdf = PdfFileReader(f)
                page_1 = pdf.getPage(0)
                if page_1.get('/Rotate', 0) in [90, 270]:
                    return page_1['/MediaBox'][2], page_1['/MediaBox'][3]
                else:
                    return page_1['/MediaBox'][3], page_1['/MediaBox'][2]
        except Exception as e:
            print(e)

    def auto_save_filename(self, file_name):
        """
        func:判断文件重名,自动在文件名后加（i）
        :param file_name:文件旧名称
        :return file_name: 文件新名称
        """
        directory, file_name = os.path.split(file_name)
        pattern = '(\d+)\)≥\.'
        while os.path.isfile(file_name):
            if re.search(pattern, file_name) is None:
                file_name = file_name.replace('.', '(0).')
            else:
                current_number = int(re.findall(pattern, file_name)[-1])
                new_number = current_number + 1
                file_name = file_name.replace(f'({current_number}).', f'({new_number}).')
            file_name = os.path.join(directory + os.sep + file_name)
        return file_name

    def file_work_arrange(self, file_name, dst_path):  # sourcery no-metrics
        """
        func: 将文件按性质移动到相应的客户文件夹的子文件夹中，此时默认文件已经属于dst_path代表的客户文件，不进行归属判断
        :param file_name: 待移动文件名
        :param dst_path: 传入客户文件夹名称
        :return: 无
        """
        file_dict = {
            '基本资料': ['执照', '章程', '开户', '身份证'],
            '财务资料': {
                '财务账表': ['财务报表', '资产负债表', '利润表', '现金流量表', '损益表', '科目余额', '明细账', '序时账', '账套', '会计准则', 'CWBB'],
                '银行流水': ['网银', '银行', '流水', '对账单', '回单', '招行', '民生', '工行', '建行', '农行', '华夏'],
                '电子发票': ['电子发票'],
            },
            '税务资料': {
                '申报表' : ['申报', '纳税', 'YBNSR', '报税', 'ZZS', 'QYSD', '工资薪金', '完税证明', '劳务报酬'],
                '进项明细': ['进项', '抵扣', '认证', '勾选'],
                '税务档案': ['申报', '纳税', 'YBNSR', '报税', 'ZZS', 'QYSD', '工资薪金', '完税证明', '劳务报酬']
            },
            '其他资料': []
        }
        try:
            for fk, fv in file_dict.items():
                move_path = Path(dst_path) / fk
                if not move_path.exists():
                    move_path.mkdir()
                if isinstance(fv, dict):  # 使用isinstance检测数据类型
                    for fvk, fvv in fv.items():
                        sun_move_path = move_path / fvk
                        if not sun_move_path.exists():
                            sun_move_path.mkdir()
                        if any(s in str(file_name) for s in fvv):
                            try:
                                Path(file_name).replace(move_path / Path(file_name).name)
                            except Exception as e:
                                # 如果是重复文件错误，调用self.auto_save_filename()方法重命名
                                if 'file already exists' in e:
                                    file_name = self.auto_save_filename(file_name)
                                    Path(file_name).replace(move_path / Path(file_name).name)
                elif isinstance(fv, list):
                    if any(s in str(file_name) for s in fv):
                        try:
                            Path(file_name).replace(move_path / Path(file_name).name)
                        except Exception as e:
                            # 如果是重复文件错误，调用self.auto_save_filename()方法重命名
                            if 'file already exists' in e:
                                file_name = self.auto_save_filename(file_name)
                                Path(file_name).replace(move_path / Path(file_name).name)
        except Exception as e:
            print(e)

    def read_from_excel(self, file, table):
        """
        :param file: excel文件的路径全称，例如：r'I:\2、work\1、FullRain_System\客户查询系统.xlsm'
        :param table: excel表格名称
        :return: 返回一个列表
        """
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        app.screen_updating = False
        wb = app.books.open(file)
        result = []
        try:
            sht = wb.sheets[table]
            a = sht.range('A1').expand().value
            for i in range(len(a)):
                row = []
                for k in range(len(a[i])):
                    if isinstance(a[i][k], datetime.datetime):
                        row.append(a[i][k].strftime("%Y-%m-%d"))
                    elif isinstance(a[i][k], float):
                        row.append(int(a[i][k]))
                    else:
                        row.append(str(a[i][k]).strip())
                result.append(row)
        finally:
            wb.close()
            app.quit()
        return result


class PathHundle:
    """
        文件夹操作类
        """

    def path_traverse(self, scr_path, all_files):
        """
        :param scr_path: 需遍历的文件夹目录
        :param all_files: 空列表，用来返回值
        :return: 返还包含所有文件名的列表
        """
        file_list = os.listdir(scr_path)
        # 准备循环判断每个元素是否是文件夹还是文件，是文件的话，把名称传入list，是文件夹的话，递归
        for file_name in file_list:
            cur_path = os.path.join(scr_path, file_name)
            if os.path.isdir(cur_path):
                self.path_traverse(cur_path, all_files)
            else:
                all_files.append(cur_path)
        return all_files

    def get_desktop_path(self):
        """获取桌面文件夹函数 """
        key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,
                             r'Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders', )
        return winreg.QueryValueEx(key, "Desktop")[0]

    def dir_mk_by_cus(self, save_base_path):
        """建立客户文件夹"""

        result = [f'{str(dt.cus_id)}-{str(dt.cus_simple_name)}-{str(dt.cus_credit_code)}' for dt in Customer.select()]

        Path(save_base_path).mkdir(exist_ok=True, parents=True)
        # 建立各客户文件夹
        p, h = 0, 0
        for cr in result:
            sun_path = Path(save_base_path) / cr
            if not sun_path.exists():
                sun_path.mkdir()
                p += 1
            for k in CUS_FILE_RANGE:
                an_path = sun_path / k
                an_path.mkdir(exist_ok=True, parents=True)
                h += 1
        print(f"客户文件夹创建成功，合计{p}个")
        print(f"子文件夹创建成功，合计{h}个")

    def is_dir_empty(self, path):
        """
        判断是否为空文件夹,如果为空，则删除
        """
        if os.path.isdir(path):
            length = len(os.listdir(path))
            if length > 0:
                print(f'注意：{path}不为空，大小{length}')
                for p in os.listdir(path):
                    pf = os.path.join(path, p)
                    self.is_dir_empty(pf)
            else:
                try:
                    os.rmdir(path)
                except Exception as e:
                    print(e)
                    self.del_with_cmd(path)
                    print(f'{path}为空文件夹, 已删除')

    def del_with_cmd(self, path):
        try:
            if os.path.isdir(path):
                cmd = 'del "' + path + '" /F'
                print(cmd)
                os.system(cmd)
        except Exception as e:
            print(e)


class FinanceHundle:
    """财务数据清洗接口的功能类"""

    def get_info_from_weixinpay(self):
        pass

    def get_info_from_bjrzbank(self):
        """将北京润梓建行账户流水处理存入finance数据表"""

        data = database.panmysql.PanMySql(database.config.FR_WORK_DB_CONFIG).select_record(table_name='bjrzbank')
        for d in data:
            try:
                use_dict = {}
                if data['']:
                    use_dict[''] = d['']

                # 将数据插入数据库
                database.panmysql.PanMySql(database.config.FR_WORK_DB_CONFIG).insert_record(
                    table_name='finance', records=use_dict)
                # 标记建行流水表的记账字段为‘已记账’
                # TODO
                # 待写修改建行流水表的的字段字典和where条件
                # PanMySql(FR_WORK_DB_CONFIG).update_record(table_name='bjrzbank', )
            except Exception as e:
                print(e)

    def get_info_from_shebao(self, excel_file: str):
        """
        将社保系统下载的明细表格数据解析成有效字段

        Args:
            excel_file (str): Excel文件的绝对路径
        """
        data_list1 = read_from_excel(file=excel_file)     # 提取的excel内容，格式为二维列表
        data_list = [['' if x is None else x for x in each_list]
                            for each_list in data_list1]
        results = []
        if '四险缴费月报人员明细过录表' in data_list[0][1]:
            for i in data_list:
                i = [x or 0 for x in i]
                try:
                    if i[0] and int(i[0]):
                        item_dict1 = {}
                        self._extracted_from_get_info_from_shebao_13(data_list, item_dict1, i, results)
                except Exception as e:
                    continue
        elif '基本医疗保险' in data_list[3][5]:
            excel_file_name = os.path.split(excel_file)[1]
            for j in data_list:
                try:
                    if int(j[1]):
                        item_dict2 = {}
                        self._extracted_from_get_info_from_shebao_34(excel_file_name, item_dict2, j, results)
                except Exception as e:
                    continue
        else:
            print('非五险明细文件')
        return results
    

class MakeTaxBook:
    """生成月度工作及申报台账"""

    def __init__(self, save_path, period):
        """类初始参数

        Args:
            save_path (str): 生成的台账excel文件保存路径
            period (str): 生成台账的所属期
        """
        self.save_path = save_path
        self.period = period

    def get_cus_info(self, table_name):
        """[从数据库获取信息，返回字典格式]

        Args:
            table_name ([str]): [数据表名称]

        Returns:
            [list]: [返回列表嵌套字典的数据集]
        """
        return PanMySql.select_record(table_name=table_name)

    def make_month_taxbook(self, info_list=None):
        """生成月度申报台账

        Args:
            info_list (list): 从数据库获取的客户信息列表，一维列表，嵌套字典形式
        """

        taxbook_dict_keys = ["客户ID", "客户名称", "统一社会信用代码", "税控托管", "纳税人类型", "收入", "销项", "进项",
                             "应纳税额", "城建税", "教育费附加", "其他地方附加", "个人所得税", "季度所得税", "文化事业建设费", "房产税",
                             "城镇土地使用税", "残保金", "印花税", "汇算清缴", "工商年报", "账", "其他"]

        workbook_dict_keys = ["客户ID", "客户名称", "统一社会信用代码", "托管项目", "纳税人类型", "税控抄报", "回单打印",
                              "粘贴发票", "收入分录", "进项分录", "费用分录", "薪酬分录", "银行分录", "折旧摊销", "税金计提结转",
                              "赤字科目调整", "结转成本", "结账", "稽核检查", "凭证打印", "凭证装订", "报表下载", "其他"]

        warning_info = ["1、代管税控客户务必抄税、清卡；", "2、比对银行发生额、余额是否一致；", "3、比对工资、社保、公积金计提发放分录与银行扣款；",
                        "4、现金、存货等流动性资产科目是否赤字；", "5、服务类企业成本结转比例是否连贯；"]

        if info_list is None:
            info_list = self.get_cus_info('customer')

        taxbook_dict = {x: "" for x in taxbook_dict_keys}
        workbook_dict = {x: "" for x in workbook_dict_keys}

        all_tax_dict = {}
        cus_list = []
        for every_cus in info_list:
            '''
            taxbook_dict['客户ID'] = every_cus['客户编号']
            taxbook_dict['客户名称'] = every_cus['客户名称']
            taxbook_dict['统一社会信用代码'] = every_cus['统一社会信用代码']
            taxbook_dict['托管项目'] = every_cus['托管项目']
            taxbook_dict['纳税人类型'] = every_cus['纳税人类型'] 
            cus_list.append(taxbook_dict)
            '''

            all_tax_dict[every_cus['服务员工']] = taxbook_dict

        print(all_tax_dict)
        return all_tax_dict

    def write_to_excel(self, info_dict: dict):
        for key, value in info_dict.items():
            if key:
                if not Path(self.save_path).exists():
                    Path(self.save_path).mkdir()

                # wb_name = f'{Path(self.save_path)} \ 申报台账-{key}.xlsx'
                wb_name = os.path.join(self.save_path, f'申报台账-{key}.xlsx')
                try:
                    wb = openpyxl.load_workbook(wb_name)
                    ws = wb[self.period] if self.period in wb.sheetnames else wb.create_sheet(title=self.period)

                except FileNotFoundError:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = self.period

                ws.merge_cells('A1:W1')
                ws['A1'] = '丰润财税月度申报台账'
                A1_style = Font(name='等线', size=16, italic=False, bold=True)
                ws['A1'].font = A1_style
                ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
                ws.row_dimensions[1].height = 28
                ws.row_dimensions[2].height = 28
                ws.row_dimensions[3].height = 28
                ws['E2'] = '所属期：'
                ws['F2'] = self.period
                ws['H2'] = '所属员工：'
                ws['I2'] = key
                for i, key_word in enumerate(value):
                    ws.cell(row=3, column=i + 1, value=key_word)
                title_style = Font(name='等线', size=11, italic=False, bold=True)
                ws['A2:W2'].font = title_style
                wb.save(wb_name)
                
                
class ParseExcel(object):
    """解析excel文件"""

    def __init__(self, filename):
        try:
            self.filename = filename
            self.__wb = openpyxl.load_workbook(self.filename)  # 打开excel
        except FileNotFoundError as e:
            raise e

    def get_max_row_num(self, sheet_name):
        """获取最大行号"""
        return self.__wb[sheet_name].max_row

    def get_max_column_num(self, sheet_name):
        """获取最大列号"""
        return self.__wb[sheet_name].max_column

    def get_cell_value(self, sheet_name, coordinate=None, row=None, column=None):
        # sourcery skip: raise-specific-error
        """获取指定单元格的数据"""
        if coordinate is not None:
            try:
                return self.__wb[sheet_name][coordinate].value
            except Exception as e:
                raise e
        elif coordinate is None and row is not None and column is not None:
            if isinstance(row, int) and isinstance(column, int):
                return self.__wb[sheet_name].cell(row=row, column=column).value
            else:
                raise TypeError('row and column must be type int')
        else:
            raise Exception("Insufficient Coordinate of cell!")

    def get_row_value(self, sheet_name, row):
        """获取某一行的数据"""
        column_num = self.get_max_column_num(sheet_name)
        if isinstance(row, int):
            row_value = []
            for column in range(1, column_num + 1):
                values_row = self.__wb[sheet_name].cell(row, column).value
                row_value.append(values_row)
            return row_value
        else:
            raise TypeError('row must be type int')

    def get_column_value(self, sheet_name, column):
        """获取某一列数据"""
        row_num = self.get_max_column_num(sheet_name)
        if isinstance(column, int):
            column_value = []
            for row in range(1, row_num + 1):
                values_column = self.__wb[sheet_name].cell(row, column).value
                column_value.append(values_column)
            return column_value
        else:
            raise TypeError('column must be type int')

    def get_all_value_1(self, sheet_name):
        """获取指定表单的所有数据(除去表头)"""
        max_row_num = self.get_max_row_num(sheet_name)
        max_column = self.get_max_column_num(sheet_name)
        values = []
        for row in range(2, max_row_num + 1):
            value_list = []
            for column in range(1, max_column + 1):
                value = self.__wb[sheet_name].cell(row, column).value
                value_list.append(value)
            values.append(value_list)
        return values

    def get_all_value_2(self, sheet_name):
        """获取指定表单的所有数据(除去表头)"""
        rows_obj = self.__wb[sheet_name].iter_rows(min_row=2, max_row=self.__wb[sheet_name].max_row, values_only=True)
        return [list(row_tuple) for row_tuple in rows_obj]

    def get_excel_title(self, sheet_name):
        """获取sheet表头"""
        return tuple(self.__wb[sheet_name].iter_rows(max_row=1, values_only=True))[0]

    def get_listdict_all_value(self, sheet_name):
        """获取所有数据，返回嵌套字典的列表"""
        sheet_title = self.get_excel_title(sheet_name)
        all_values = self.get_all_value_2(sheet_name)
        return [dict(zip(sheet_title, value)) for value in all_values]

    def get_list_nametuple_all_value(self, sheet_name):
        """获取所有数据，返回嵌套命名元组的列表"""
        sheet_title = self.get_excel_title(sheet_name)
        values = self.get_all_value_2(sheet_name)
        excel = namedtuple('excel', sheet_title)
        value_list = []
        for value in values:
            e = excel(*value)
            value_list.append(e)
        return value_list

    def write_cell(self, sheet_name, row, column, value=None, bold=True, color=BLACK):
        if isinstance(row, int) and isinstance(column, int):
            try:
                cell_obj = self.__wb[sheet_name].cell(row, column)
                cell_obj.font = Font(color=color, bold=bold)
                cell_obj.value = value
                self.__wb.save(self.filename)
            except Exception as e:
                raise e
        else:
            raise TypeError('row and column must be type int')
        
        
class PayRoll:
    """
    生成月工资表类
    """
    PERSON_PAYROLL={}
    
    def __init__(self, the_month:str):
        """
        类的初始参数    
        Args:
            the_month (str): 指定生成月工资表的所属月份，例如‘202201’
        """
        self.the_month = datetime.strptime(the_month, '%Y-%m-%d')
        
    def get_base_salary(self, base_salary=3000):
        """
        根据工龄计算员工基本工资，每满一年上浮10%

        Args:
            the_date (_type_): 截止的时间
            base_salary (int, optional): 基本工资起始标准，默认3000.

        Returns:
            list: 包含多个字典的列表
        """
        query = (User
                .select()
                .where(User.person_state == '在职'))
        # the_date = datetime.strptime(self.the_month, '%Y-%m-%d')
        for user in query:
            base_salary = user.basic_salary or 3000
            work_year_count = get_differ_month(user.entry_date, self.the_month) // 12
            for _ in range(work_year_count):
                base_salary = round(base_salary * 0.1 + base_salary, 2)
            user_base_salary = {user.name: {'base_salary': base_salary}}
            self.PERSON_PAYROLL.append(user_base_salary)

    def get_meritpay(self):
        """
        计算绩效工资，主要是根据客户表中月服务费字段计算每个人的客户服务总额
        """
        query = (Customer
                 .select(Customer.service_personnel.alias('姓名'),
                        fn.SUM(Customer.month_pay).alias('服务费总额'))
                 .join(User)
                 .where((Customer.cus_state == '正常') | (Customer.cus_state == '无需申报') &
                        (User.person_state == '在职') & 
                        (Customer.start_service_date <= self.the_month))
                 .group_by(Customer.service_personnel))
        for user in query.objects():
            for key in self.PERSON_PAYROLL:
                if user.name == key:
                    self.PERSON_PAYROLL[user.name]['绩效工资'] = user.服务费总额 * 0.1

    def get_shebao_data(self):
        """
        获取五险一金数据
        """
        query = (SocialHouse
                 .select()
                 .where((SocialHouse.social_year == self.the_month.year) & 
                        (SocialHouse.social_month == self.the_month.month)))
        for user in query.objects():
            for key in self.PERSON_PAYROLL:
                if user.person == key:
                    self.PERSON_PAYROLL[user.person]['养老'] = user.yanglao_person
                    self.PERSON_PAYROLL[user.person]['失业'] = user.shiye_person
                    self.PERSON_PAYROLL[user.person]['医疗'] = user.yiliao_person
                    self.PERSON_PAYROLL[user.person]['公积金'] = user.house_fund_person
                    
    def get_bonus(self):
        """
        计算项目奖金，主要是根据project表中当期已完成项目未发放的奖金
        """
        pass
    
    def run_func(self):
        self.get_base_salary()
        self.get_meritpay()
        self.get_bonus()
        self.get_shebao_data()
        return PERSON_PAYROLL
    